"""Repositorio SQL (SQLite) para gerir o inventario."""

from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO
from pathlib import Path
import re
import sqlite3
from threading import RLock
from typing import Dict, List, Optional
import unicodedata
from uuid import uuid4

from openpyxl import Workbook
from app.core.config import LEGACY_EXCEL_PATH, LEGACY_EXCEL_SHEET_NAME, SQLITE_DB_PATH
from app.schemas import (
    ColunaSchema,
    DashboardResumoVendas,
    DashboardSerieValor,
    DashboardVendas,
    Peca,
    PecaCreate,
    PecaUpdate,
    RegistoVendaResultado,
    VendaHistoricoItem,
)


class ExcelRepository:
    """
    Repositorio principal do inventario.

    Nota: o nome da classe foi mantido para evitar mexidas adicionais no resto da app.
    A persistencia agora e feita em SQLite.
    """

    def __init__(self) -> None:
        self._lock = RLock()
        self._core_keys = {"referencia", "categoria", "marca", "designacao", "preco", "quantidade", "local", "id"}
        self._fixed_core_keys = {"preco", "quantidade", "id"}
        self._core_columns = [
            ColunaSchema(chave="referencia", nome="Referência", removivel=True),
            ColunaSchema(chave="categoria", nome="Categoria", removivel=True),
            ColunaSchema(chave="marca", nome="Marca", removivel=True),
            ColunaSchema(chave="designacao", nome="Designação", removivel=True),
            ColunaSchema(chave="preco", nome="Preço", removivel=False),
            ColunaSchema(chave="quantidade", nome="Quantidade", removivel=False),
            ColunaSchema(chave="local", nome="Local", removivel=True),
        ]

        self._validar_caminho_bd()
        self._inicializar_schema()
        self._migrar_excel_legacy_se_necessario()

    def listar(self, termo: Optional[str] = None) -> List[Peca]:
        """Lista pecas com filtro opcional por texto."""
        with self._lock, self._connect() as conn:
            rows = conn.execute(
                """
                SELECT id, referencia, categoria, marca, designacao, preco, quantidade, local, extras
                FROM pecas
                ORDER BY LOWER(designacao), LOWER(referencia)
                """
            ).fetchall()

            pecas = [self._row_to_peca(row) for row in rows]
            if not termo:
                return pecas

            termo_normalizado = termo.strip().lower()
            return [peca for peca in pecas if self._match_termo(peca, termo_normalizado)]

    def listar_vendas(self) -> List[VendaHistoricoItem]:
        """Lista o historico de vendas mais recentes."""
        with self._lock, self._connect() as conn:
            rows = conn.execute(
                """
                SELECT id, peca_id, referencia, categoria, marca, designacao, local,
                       preco_unitario, quantidade_vendida, vendida_em, extras
                FROM vendas
                ORDER BY vendida_em DESC, id DESC
                """
            ).fetchall()
            return [self._row_to_venda(row) for row in rows]

    def atualizar_venda_historico(
        self,
        venda_id: str,
        quantidade_vendida: int,
        preco_unitario: float,
    ) -> Optional[VendaHistoricoItem]:
        """Edita uma venda antiga e ajusta o stock atual de acordo com a diferenca de quantidade."""
        quantidade_vendida = max(1, int(quantidade_vendida))
        preco_unitario = self._safe_price(preco_unitario)
        if preco_unitario < 0:
            raise ValueError("Indica um preco unitario valido.")

        with self._lock, self._connect() as conn:
            venda_row = self._obter_venda_por_id_conn(conn, venda_id)
            if not venda_row:
                return None

            venda = self._row_to_venda(venda_row)
            delta_quantidade = quantidade_vendida - max(1, int(venda.quantidade_vendida))
            if delta_quantidade > 0:
                self._retirar_do_stock(conn, venda, delta_quantidade)
            elif delta_quantidade < 0:
                self._repor_ao_stock(conn, venda, abs(delta_quantidade))

            conn.execute(
                """
                UPDATE vendas
                SET quantidade_vendida = ?, preco_unitario = ?
                WHERE id = ?
                """,
                (quantidade_vendida, preco_unitario, venda_id),
            )
            conn.commit()

            venda_row_atualizada = self._obter_venda_por_id_conn(conn, venda_id)
            return self._row_to_venda(venda_row_atualizada) if venda_row_atualizada else None

    def eliminar_venda_historico(self, venda_id: str) -> bool:
        """Remove uma venda do historico sem mexer no stock."""
        with self._lock, self._connect() as conn:
            deleted = conn.execute("DELETE FROM vendas WHERE id = ?", (venda_id,))
            conn.commit()
            return deleted.rowcount > 0

    def repor_venda_no_inventario(self, venda_id: str) -> Optional[Peca]:
        """Repõe no stock a quantidade vendida e remove o respetivo registo do historico."""
        with self._lock, self._connect() as conn:
            venda_row = self._obter_venda_por_id_conn(conn, venda_id)
            if not venda_row:
                return None

            venda = self._row_to_venda(venda_row)
            peca = self._repor_ao_stock(conn, venda, max(1, int(venda.quantidade_vendida)))
            conn.execute("DELETE FROM vendas WHERE id = ?", (venda_id,))
            conn.commit()
            return peca

    def obter_dashboard_vendas(self) -> DashboardVendas:
        """Constroi o dashboard com historico, faturacao e valor atual em stock."""
        historico = self.listar_vendas()
        pecas = self.listar()

        faturacao_total = round(sum(venda.total_venda for venda in historico), 2)
        valor_em_stock = round(sum(self._safe_price(peca.preco) * max(0, int(peca.quantidade)) for peca in pecas), 2)
        total_vendas = len(historico)
        unidades_vendidas = sum(max(0, int(venda.quantidade_vendida)) for venda in historico)
        unidades_em_stock = sum(max(0, int(peca.quantidade)) for peca in pecas)

        faturacao_por_mes_map: Dict[str, float] = {}
        for venda in historico:
            chave = venda.vendida_em.strftime("%Y-%m")
            faturacao_por_mes_map[chave] = round(faturacao_por_mes_map.get(chave, 0) + venda.total_venda, 2)

        faturacao_por_mes = [
            DashboardSerieValor(etiqueta=self._formatar_mes_dashboard(chave), valor=valor)
            for chave, valor in sorted(faturacao_por_mes_map.items())
        ]

        valor_stock_por_categoria_map: Dict[str, float] = {}
        for peca in pecas:
            categoria = str(peca.categoria or "").strip() or "Sem categoria"
            valor_peca = round(self._safe_price(peca.preco) * max(0, int(peca.quantidade)), 2)
            if valor_peca <= 0:
                continue
            valor_stock_por_categoria_map[categoria] = round(
                valor_stock_por_categoria_map.get(categoria, 0) + valor_peca,
                2,
            )

        valor_stock_por_categoria = [
            DashboardSerieValor(etiqueta=etiqueta, valor=valor)
            for etiqueta, valor in sorted(
                valor_stock_por_categoria_map.items(),
                key=lambda item: (-item[1], item[0].lower()),
            )[:8]
        ]

        faturacao_por_material_map: Dict[str, float] = {}
        for venda in historico:
            etiqueta = self._montar_etiqueta_material(venda.referencia, venda.designacao)
            faturacao_por_material_map[etiqueta] = round(
                faturacao_por_material_map.get(etiqueta, 0) + venda.total_venda,
                2,
            )

        faturacao_por_material = [
            DashboardSerieValor(etiqueta=etiqueta, valor=valor)
            for etiqueta, valor in sorted(
                faturacao_por_material_map.items(),
                key=lambda item: (-item[1], item[0].lower()),
            )[:8]
        ]

        return DashboardVendas(
            resumo=DashboardResumoVendas(
                faturacao_total=faturacao_total,
                valor_em_stock=valor_em_stock,
                total_vendas=total_vendas,
                unidades_vendidas=unidades_vendidas,
                unidades_em_stock=unidades_em_stock,
            ),
            faturacao_por_mes=faturacao_por_mes,
            valor_stock_por_categoria=valor_stock_por_categoria,
            faturacao_por_material=faturacao_por_material,
            historico=historico,
        )

    def obter_por_id(self, peca_id: str) -> Optional[Peca]:
        """Procura uma peca pelo ID."""
        with self._lock, self._connect() as conn:
            row = conn.execute(
                """
                SELECT id, referencia, categoria, marca, designacao, preco, quantidade, local, extras
                FROM pecas
                WHERE id = ?
                """,
                (peca_id,),
            ).fetchone()

            return self._row_to_peca(row) if row else None

    def criar(self, payload: PecaCreate) -> Peca:
        """Cria uma nova peca."""
        with self._lock, self._connect() as conn:
            peca_id = str(uuid4())
            extras_validos = self._filtrar_extras_validos(conn, payload.extras)

            conn.execute(
                """
                INSERT INTO pecas (id, referencia, categoria, marca, designacao, preco, quantidade, local, extras)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    peca_id,
                    payload.referencia,
                    payload.categoria,
                    payload.marca,
                    payload.designacao,
                    payload.preco,
                    payload.quantidade,
                    payload.local,
                    json.dumps(extras_validos, ensure_ascii=False),
                ),
            )
            conn.commit()

            return Peca(
                id=peca_id,
                referencia=payload.referencia,
                categoria=payload.categoria,
                marca=payload.marca,
                designacao=payload.designacao,
                preco=payload.preco,
                quantidade=payload.quantidade,
                local=payload.local,
                extras=extras_validos,
            )

    def atualizar(self, peca_id: str, payload: PecaUpdate) -> Optional[Peca]:
        """Atualiza todos os campos de uma peca."""
        with self._lock, self._connect() as conn:
            existe = conn.execute("SELECT id FROM pecas WHERE id = ?", (peca_id,)).fetchone()
            if not existe:
                return None

            extras_validos = self._filtrar_extras_validos(conn, payload.extras)
            conn.execute(
                """
                UPDATE pecas
                SET referencia = ?,
                    categoria = ?,
                    marca = ?,
                    designacao = ?,
                    preco = ?,
                    quantidade = ?,
                    local = ?,
                    extras = ?
                WHERE id = ?
                """,
                (
                    payload.referencia,
                    payload.categoria,
                    payload.marca,
                    payload.designacao,
                    payload.preco,
                    payload.quantidade,
                    payload.local,
                    json.dumps(extras_validos, ensure_ascii=False),
                    peca_id,
                ),
            )
            conn.commit()

            return Peca(
                id=peca_id,
                referencia=payload.referencia,
                categoria=payload.categoria,
                marca=payload.marca,
                designacao=payload.designacao,
                preco=payload.preco,
                quantidade=payload.quantidade,
                local=payload.local,
                extras=extras_validos,
            )

    def atualizar_quantidade(self, peca_id: str, nova_quantidade: int) -> Optional[Peca]:
        """Atualiza apenas a quantidade de uma peca."""
        with self._lock, self._connect() as conn:
            updated = conn.execute(
                "UPDATE pecas SET quantidade = ? WHERE id = ?",
                (max(0, nova_quantidade), peca_id),
            )
            conn.commit()
            if updated.rowcount == 0:
                return None

            row = conn.execute(
                """
                SELECT id, referencia, categoria, marca, designacao, preco, quantidade, local, extras
                FROM pecas
                WHERE id = ?
                """,
                (peca_id,),
            ).fetchone()
            return self._row_to_peca(row) if row else None

    def registar_venda(self, peca_id: str, quantidade_vendida: int, preco_unitario: float) -> Optional[RegistoVendaResultado]:
        """Regista uma venda e reduz o stock da peca."""
        quantidade_vendida = max(0, int(quantidade_vendida))
        if quantidade_vendida <= 0:
            raise ValueError("Indica uma quantidade valida para a venda.")
        preco_unitario = self._safe_price(preco_unitario)
        if preco_unitario < 0:
            raise ValueError("Indica um preco valido para a venda.")

        with self._lock, self._connect() as conn:
            row = conn.execute(
                """
                SELECT id, referencia, categoria, marca, designacao, preco, quantidade, local, extras
                FROM pecas
                WHERE id = ?
                """,
                (peca_id,),
            ).fetchone()
            if not row:
                return None

            peca = self._row_to_peca(row)
            stock_atual = max(0, int(peca.quantidade))
            if quantidade_vendida > stock_atual:
                raise ValueError("A quantidade vendida nao pode ser superior ao stock atual.")

            self._criar_registo_venda(conn, peca, quantidade_vendida, preco_unitario)

            quantidade_restante = stock_atual - quantidade_vendida
            if quantidade_restante <= 0:
                conn.execute("DELETE FROM pecas WHERE id = ?", (peca_id,))
                conn.commit()
                return RegistoVendaResultado(
                    removida_do_inventario=True,
                    quantidade_restante=0,
                    peca=None,
                )

            conn.execute(
                "UPDATE pecas SET quantidade = ? WHERE id = ?",
                (quantidade_restante, peca_id),
            )
            conn.commit()

            return RegistoVendaResultado(
                removida_do_inventario=False,
                quantidade_restante=quantidade_restante,
                peca=Peca(
                    id=peca.id,
                    referencia=peca.referencia,
                    categoria=peca.categoria,
                    marca=peca.marca,
                    designacao=peca.designacao,
                    preco=peca.preco,
                    quantidade=quantidade_restante,
                    local=peca.local,
                    extras=peca.extras,
                ),
            )

    def eliminar(self, peca_id: str) -> bool:
        """Elimina uma peca pelo ID."""
        with self._lock, self._connect() as conn:
            deleted = conn.execute("DELETE FROM pecas WHERE id = ?", (peca_id,))
            conn.commit()
            return deleted.rowcount > 0

    def obter_colunas(self) -> List[ColunaSchema]:
        """Lista colunas visiveis no schema."""
        with self._lock, self._connect() as conn:
            core_names_rows = conn.execute(
                """
                SELECT chave, nome
                FROM schema_core_nomes
                """
            ).fetchall()
            core_name_map = {row["chave"]: row["nome"] for row in core_names_rows}
            ordem_rows = conn.execute(
                """
                SELECT chave, posicao
                FROM schema_ordem_colunas
                ORDER BY posicao ASC
                """
            ).fetchall()
            ordem = {row["chave"]: int(row["posicao"]) for row in ordem_rows}
            ordem_definida = bool(ordem)

            colunas_core = [
                ColunaSchema(
                    chave=coluna.chave,
                    nome=core_name_map.get(coluna.chave, coluna.nome),
                    removivel=coluna.removivel,
                )
                for coluna in self._core_columns
                if not ordem_definida or coluna.chave in ordem or coluna.chave in self._fixed_core_keys
            ]

            rows = conn.execute(
                """
                SELECT chave, nome
                FROM schema_colunas
                ORDER BY posicao ASC
                """
            ).fetchall()
            dinamicas = [ColunaSchema(chave=row["chave"], nome=row["nome"], removivel=True) for row in rows]
            todas = [*colunas_core, *dinamicas]

            todas.sort(key=lambda coluna: ordem.get(coluna.chave, 9999))
            return self._normalizar_ordem_colunas(todas)

    def adicionar_coluna(self, nome: str) -> ColunaSchema:
        """Adiciona uma nova coluna dinamica."""
        with self._lock, self._connect() as conn:
            nome_limpo = nome.strip()
            chave = self._normalizar_chave_coluna(nome_limpo)
            if not chave:
                raise ValueError("O nome da coluna nao e valido.")
            if chave in self._core_keys:
                if chave in self._fixed_core_keys:
                    raise ValueError("Ja existe uma coluna com esse nome.")

                coluna_core = next((coluna for coluna in self._core_columns if coluna.chave == chave), None)
                if not coluna_core:
                    raise ValueError("Ja existe uma coluna com esse nome.")

                visivel = conn.execute(
                    "SELECT 1 FROM schema_ordem_colunas WHERE chave = ?",
                    (chave,),
                ).fetchone()
                if visivel:
                    raise ValueError("Ja existe uma coluna com esse nome.")

                proxima_pos = conn.execute(
                    "SELECT COALESCE(MAX(posicao), 0) AS max_pos FROM schema_ordem_colunas"
                ).fetchone()
                conn.execute(
                    """
                    INSERT INTO schema_core_nomes (chave, nome)
                    VALUES (?, ?)
                    ON CONFLICT(chave) DO UPDATE SET nome = excluded.nome
                    """,
                    (chave, nome_limpo),
                )
                conn.execute(
                    "INSERT OR REPLACE INTO schema_ordem_colunas (chave, posicao) VALUES (?, ?)",
                    (chave, int(proxima_pos["max_pos"]) + 1),
                )
                conn.commit()
                return ColunaSchema(chave=chave, nome=nome_limpo, removivel=coluna_core.removivel)

            existente = conn.execute("SELECT 1 FROM schema_colunas WHERE chave = ?", (chave,)).fetchone()
            if existente:
                raise ValueError("Ja existe uma coluna com esse nome.")

            posicao_row = conn.execute("SELECT COALESCE(MAX(posicao), 0) AS max_pos FROM schema_colunas").fetchone()
            posicao = int(posicao_row["max_pos"]) + 1

            conn.execute(
                "INSERT INTO schema_colunas (chave, nome, posicao) VALUES (?, ?, ?)",
                (chave, nome_limpo, posicao),
            )
            proxima_pos = conn.execute(
                "SELECT COALESCE(MAX(posicao), 0) AS max_pos FROM schema_ordem_colunas"
            ).fetchone()
            conn.execute(
                "INSERT OR REPLACE INTO schema_ordem_colunas (chave, posicao) VALUES (?, ?)",
                (chave, int(proxima_pos["max_pos"]) + 1),
            )
            conn.commit()
            return ColunaSchema(chave=chave, nome=nome_limpo, removivel=True)

    def remover_coluna(self, chave: str) -> bool:
        """Remove uma coluna dinamica e limpa os dados antigos em extras."""
        with self._lock, self._connect() as conn:
            chave_normalizada = self._normalizar_chave_coluna(chave)
            if not chave_normalizada:
                return False
            if chave_normalizada in self._fixed_core_keys:
                raise ValueError("Nao e permitido remover colunas fixas.")
            if chave_normalizada in self._core_keys:
                existe_core = any(coluna.chave == chave_normalizada for coluna in self._core_columns)
                if not existe_core:
                    return False

                removed = conn.execute("DELETE FROM schema_ordem_colunas WHERE chave = ?", (chave_normalizada,))
                conn.commit()
                return removed.rowcount > 0

            existente = conn.execute("SELECT 1 FROM schema_colunas WHERE chave = ?", (chave_normalizada,)).fetchone()
            if not existente:
                return False

            conn.execute("DELETE FROM schema_colunas WHERE chave = ?", (chave_normalizada,))
            conn.execute("DELETE FROM schema_ordem_colunas WHERE chave = ?", (chave_normalizada,))

            rows = conn.execute("SELECT id, extras FROM pecas").fetchall()
            for row in rows:
                extras = self._parse_extras(row["extras"])
                if chave_normalizada not in extras:
                    continue
                extras.pop(chave_normalizada, None)
                conn.execute(
                    "UPDATE pecas SET extras = ? WHERE id = ?",
                    (json.dumps(extras, ensure_ascii=False), row["id"]),
                )

            conn.commit()
            return True

    def renomear_coluna(self, chave: str, novo_nome: str) -> Optional[ColunaSchema]:
        """Renomeia uma coluna base ou dinamica."""
        with self._lock, self._connect() as conn:
            chave_normalizada = self._normalizar_chave_coluna(chave)
            nome_limpo = novo_nome.strip()
            if not chave_normalizada or not nome_limpo:
                return None

            if chave_normalizada in self._core_keys:
                if chave_normalizada == "id":
                    return None

                updated = conn.execute(
                    "UPDATE schema_core_nomes SET nome = ? WHERE chave = ?",
                    (nome_limpo, chave_normalizada),
                )
                conn.commit()
                if updated.rowcount == 0:
                    return None
                return ColunaSchema(
                    chave=chave_normalizada,
                    nome=nome_limpo,
                    removivel=chave_normalizada not in self._fixed_core_keys,
                )

            updated = conn.execute(
                "UPDATE schema_colunas SET nome = ? WHERE chave = ?",
                (nome_limpo, chave_normalizada),
            )
            conn.commit()
            if updated.rowcount == 0:
                return None
            return ColunaSchema(chave=chave_normalizada, nome=nome_limpo, removivel=True)

    def reordenar_colunas(self, chaves: List[str]) -> List[ColunaSchema]:
        """Atualiza a ordem das colunas visiveis e devolve o schema final."""
        with self._lock, self._connect() as conn:
            atuais = self.obter_colunas()
            chaves_atuais = [coluna.chave for coluna in atuais]
            chaves_requisicao = [self._normalizar_chave_coluna(chave) for chave in chaves]

            if set(chaves_requisicao) != set(chaves_atuais):
                raise ValueError("A ordem enviada deve conter exatamente as colunas atuais.")

            for posicao, chave in enumerate(chaves_requisicao, start=1):
                conn.execute(
                    """
                    INSERT INTO schema_ordem_colunas (chave, posicao)
                    VALUES (?, ?)
                    ON CONFLICT(chave) DO UPDATE SET posicao = excluded.posicao
                    """,
                    (chave, posicao),
                )
            conn.commit()
            return self.obter_colunas()

    def exportar_para_excel(self) -> bytes:
        """Exporta o estado atual da base de dados para ficheiro Excel em memoria."""
        colunas = self.obter_colunas()
        pecas = self.listar()

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"

        # Header
        ws.append([coluna.nome for coluna in colunas])
        indice_preco = next((idx for idx, coluna in enumerate(colunas, start=1) if coluna.chave == "preco"), None)

        for peca in pecas:
            linha = []
            for coluna in colunas:
                if coluna.chave in self._core_keys:
                    linha.append(getattr(peca, coluna.chave, None))
                else:
                    linha.append(peca.extras.get(coluna.chave))
            ws.append(linha)
            if indice_preco:
                ws.cell(row=ws.max_row, column=indice_preco).number_format = '#,##0.00 €'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _connect(self) -> sqlite3.Connection:
        """Cria conexao SQLite com row factory em dicionario."""
        conn = sqlite3.connect(SQLITE_DB_PATH)
        conn.row_factory = sqlite3.Row
        return conn

    def _validar_caminho_bd(self) -> None:
        """Garante que a pasta da base de dados existe."""
        Path(SQLITE_DB_PATH).parent.mkdir(parents=True, exist_ok=True)

    def _inicializar_schema(self) -> None:
        """Cria tabelas base se necessario."""
        with self._lock, self._connect() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS pecas (
                    id TEXT PRIMARY KEY,
                    referencia TEXT NOT NULL,
                    categoria TEXT NOT NULL,
                    marca TEXT NOT NULL,
                    designacao TEXT NOT NULL,
                    preco REAL NOT NULL DEFAULT 0,
                    quantidade INTEGER NOT NULL DEFAULT 0,
                    local TEXT,
                    extras TEXT NOT NULL DEFAULT '{}'
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS schema_colunas (
                    chave TEXT PRIMARY KEY,
                    nome TEXT NOT NULL,
                    posicao INTEGER NOT NULL
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS schema_core_nomes (
                    chave TEXT PRIMARY KEY,
                    nome TEXT NOT NULL
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS schema_ordem_colunas (
                    chave TEXT PRIMARY KEY,
                    posicao INTEGER NOT NULL
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS vendas (
                    id TEXT PRIMARY KEY,
                    peca_id TEXT,
                    referencia TEXT NOT NULL,
                    categoria TEXT NOT NULL,
                    marca TEXT NOT NULL,
                    designacao TEXT NOT NULL,
                    preco_unitario REAL NOT NULL DEFAULT 0,
                    quantidade_vendida INTEGER NOT NULL DEFAULT 1,
                    local TEXT,
                    vendida_em TEXT NOT NULL,
                    extras TEXT NOT NULL DEFAULT '{}'
                )
                """
            )

            for coluna in self._core_columns:
                conn.execute(
                    """
                    INSERT OR IGNORE INTO schema_core_nomes (chave, nome)
                    VALUES (?, ?)
                """,
                (coluna.chave, coluna.nome),
            )

            self._migrar_preco_core(conn)

            ordem_existente = conn.execute("SELECT COUNT(*) AS total FROM schema_ordem_colunas").fetchone()
            if int(ordem_existente["total"]) == 0:
                posicao = 1
                for coluna in self._core_columns:
                    conn.execute(
                        "INSERT INTO schema_ordem_colunas (chave, posicao) VALUES (?, ?)",
                        (coluna.chave, posicao),
                    )
                    posicao += 1

                dinamicas = conn.execute(
                    "SELECT chave FROM schema_colunas ORDER BY posicao ASC"
                ).fetchall()
                for row in dinamicas:
                    conn.execute(
                        "INSERT OR IGNORE INTO schema_ordem_colunas (chave, posicao) VALUES (?, ?)",
                        (row["chave"], posicao),
                    )
                    posicao += 1

            posicao_max = conn.execute("SELECT COALESCE(MAX(posicao), 0) AS max_pos FROM schema_ordem_colunas").fetchone()
            proxima_pos = int(posicao_max["max_pos"]) + 1
            chaves_ordenadas = {
                row["chave"] for row in conn.execute("SELECT chave FROM schema_ordem_colunas").fetchall()
            }
            chaves_atuais = [coluna.chave for coluna in self._core_columns]
            chaves_atuais.extend(
                row["chave"] for row in conn.execute("SELECT chave FROM schema_colunas ORDER BY posicao ASC").fetchall()
            )
            for chave in chaves_atuais:
                if chave in chaves_ordenadas:
                    continue
                if chave in self._core_keys and chave not in self._fixed_core_keys:
                    continue
                conn.execute(
                    "INSERT INTO schema_ordem_colunas (chave, posicao) VALUES (?, ?)",
                    (chave, proxima_pos),
                )
                proxima_pos += 1
            conn.commit()

    def _migrar_preco_core(self, conn: sqlite3.Connection) -> None:
        """Garante que o preco existe como coluna base e remove duplicados dinamicos."""
        colunas_pecas = {
            row["name"]
            for row in conn.execute("PRAGMA table_info(pecas)").fetchall()
        }
        if "preco" not in colunas_pecas:
            conn.execute("ALTER TABLE pecas ADD COLUMN preco REAL NOT NULL DEFAULT 0")

        coluna_dinamica_preco = conn.execute(
            "SELECT 1 FROM schema_colunas WHERE chave = ?",
            ("preco",),
        ).fetchone()
        if coluna_dinamica_preco:
            rows = conn.execute("SELECT id, preco, extras FROM pecas").fetchall()
            for row in rows:
                extras = self._parse_extras(row["extras"])
                preco_extra = self._safe_price(extras.pop("preco", None))
                preco_atual = self._safe_price(row["preco"])
                novo_preco = preco_atual if preco_atual > 0 else preco_extra
                conn.execute(
                    "UPDATE pecas SET preco = ?, extras = ? WHERE id = ?",
                    (novo_preco, json.dumps(extras, ensure_ascii=False), row["id"]),
                )

            conn.execute("DELETE FROM schema_colunas WHERE chave = ?", ("preco",))
            conn.execute("DELETE FROM schema_ordem_colunas WHERE chave = ?", ("preco",))

        conn.execute(
            """
            INSERT OR IGNORE INTO schema_core_nomes (chave, nome)
            VALUES (?, ?)
            """,
            ("preco", "Preço"),
        )

    def _migrar_excel_legacy_se_necessario(self) -> None:
        """
        Migra os dados do Excel antigo para SQLite uma unica vez.

        A migracao ocorre apenas se:
        - o ficheiro legacy existir
        - a tabela SQL estiver vazia
        """
        with self._lock, self._connect() as conn:
            total = conn.execute("SELECT COUNT(*) AS total FROM pecas").fetchone()["total"]
            if total > 0:
                return

        if not Path(LEGACY_EXCEL_PATH).exists():
            return

        try:
            from openpyxl import load_workbook
        except Exception:
            # Sem openpyxl nao ha migracao automatica.
            return

        wb = load_workbook(LEGACY_EXCEL_PATH)
        sheet = wb[LEGACY_EXCEL_SHEET_NAME] if LEGACY_EXCEL_SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]
        headers = [sheet.cell(row=1, column=idx).value for idx in range(1, sheet.max_column + 1)]

        key_map: Dict[int, str] = {}
        custom_headers: List[tuple[str, str]] = []
        for idx, header in enumerate(headers, start=1):
            nome = str(header or "").strip()
            if not nome:
                continue

            normalized = self._normalizar_chave_coluna(nome)
            if normalized in self._core_keys:
                key_map[idx] = normalized
                continue

            aliases = {
                "referencia": {"referencia", "ref"},
                "designacao": {"designacao"},
                "preco": {"preco", "preco_eur", "preco_euro", "preco_unitario"},
            }
            encontrado = None
            norm_no_accents = self._normalizar_texto(nome)
            for core_key, options in aliases.items():
                if norm_no_accents in options:
                    encontrado = core_key
                    break

            if encontrado and encontrado not in key_map.values():
                key_map[idx] = encontrado
                continue

            if not normalized:
                continue
            unique_key = normalized
            suffix = 2
            while unique_key in self._core_keys or any(existing[0] == unique_key for existing in custom_headers):
                unique_key = f"{normalized}_{suffix}"
                suffix += 1

            key_map[idx] = unique_key
            custom_headers.append((unique_key, nome))

        required = {"referencia", "categoria", "marca", "designacao", "quantidade"}
        if not required.issubset(set(key_map.values())):
            return

        with self._lock, self._connect() as conn:
            # Colunas dinamicas.
            pos = 1
            for key, nome in custom_headers:
                conn.execute(
                    "INSERT OR IGNORE INTO schema_colunas (chave, nome, posicao) VALUES (?, ?, ?)",
                    (key, nome, pos),
                )
                pos += 1

            # Linhas.
            for row_idx in range(2, sheet.max_row + 1):
                linha = {key_map[col]: sheet.cell(row=row_idx, column=col).value for col in key_map}
                referencia = str(linha.get("referencia") or "").strip()
                categoria = str(linha.get("categoria") or "").strip()
                marca = str(linha.get("marca") or "").strip()
                designacao = str(linha.get("designacao") or "").strip()
                if not all([referencia, categoria, marca, designacao]):
                    continue

                quantidade = self._safe_int(linha.get("quantidade"))
                preco = self._safe_price(linha.get("preco"))
                local_raw = linha.get("local")
                local = str(local_raw).strip() if local_raw is not None and str(local_raw).strip() else None
                peca_id = str(linha.get("id") or "").strip() or str(uuid4())

                extras = {}
                for key, _nome in custom_headers:
                    valor = linha.get(key)
                    if valor is None:
                        continue
                    if isinstance(valor, str):
                        valor = valor.strip()
                        if not valor:
                            continue
                    extras[key] = valor

                conn.execute(
                    """
                    INSERT OR REPLACE INTO pecas (id, referencia, categoria, marca, designacao, preco, quantidade, local, extras)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        peca_id,
                        referencia,
                        categoria,
                        marca,
                        designacao,
                        preco,
                        max(0, quantidade),
                        local,
                        json.dumps(extras, ensure_ascii=False),
                    ),
                )

            conn.commit()

    def _filtrar_extras_validos(self, conn: sqlite3.Connection, extras: dict[str, object]) -> dict[str, object]:
        """Mantem apenas extras que existem no schema dinamico."""
        allowed = {
            row["chave"]
            for row in conn.execute("SELECT chave FROM schema_colunas").fetchall()
        }

        cleaned: dict[str, object] = {}
        for key, value in (extras or {}).items():
            key_norm = self._normalizar_chave_coluna(key)
            if key_norm not in allowed:
                continue
            if value is None:
                continue
            if isinstance(value, str):
                text = value.strip()
                if not text:
                    continue
                cleaned[key_norm] = text
                continue
            cleaned[key_norm] = value
        return cleaned

    def _row_to_peca(self, row: sqlite3.Row) -> Peca:
        """Converte linha SQL para modelo Pydantic."""
        return Peca(
            id=row["id"],
            referencia=row["referencia"],
            categoria=row["categoria"],
            marca=row["marca"],
            designacao=row["designacao"],
            preco=self._safe_price(row["preco"]),
            quantidade=max(0, int(row["quantidade"])),
            local=row["local"],
            extras=self._parse_extras(row["extras"]),
        )

    def _row_to_venda(self, row: sqlite3.Row) -> VendaHistoricoItem:
        """Converte linha SQL de venda num item do historico."""
        preco_unitario = self._safe_price(row["preco_unitario"])
        quantidade_vendida = max(1, int(row["quantidade_vendida"]))
        total_venda = round(preco_unitario * quantidade_vendida, 2)
        vendida_em_raw = str(row["vendida_em"] or "").strip()
        try:
            vendida_em = datetime.fromisoformat(vendida_em_raw)
        except ValueError:
            vendida_em = datetime.now()

        return VendaHistoricoItem(
            id=row["id"],
            peca_id=row["peca_id"],
            referencia=row["referencia"],
            categoria=row["categoria"],
            marca=row["marca"],
            designacao=row["designacao"],
            local=row["local"],
            preco_unitario=preco_unitario,
            quantidade_vendida=quantidade_vendida,
            total_venda=total_venda,
            vendida_em=vendida_em,
            extras=self._parse_extras(row["extras"]),
        )

    def _obter_venda_por_id_conn(self, conn: sqlite3.Connection, venda_id: str) -> Optional[sqlite3.Row]:
        """Le um registo de venda pelo ID usando uma conexao ja aberta."""
        return conn.execute(
            """
            SELECT id, peca_id, referencia, categoria, marca, designacao, local,
                   preco_unitario, quantidade_vendida, vendida_em, extras
            FROM vendas
            WHERE id = ?
            """,
            (venda_id,),
        ).fetchone()

    def _obter_peca_por_id_conn(self, conn: sqlite3.Connection, peca_id: str) -> Optional[Peca]:
        """Le uma peca pelo ID usando uma conexao ja aberta."""
        row = conn.execute(
            """
            SELECT id, referencia, categoria, marca, designacao, preco, quantidade, local, extras
            FROM pecas
            WHERE id = ?
            """,
            (peca_id,),
        ).fetchone()
        return self._row_to_peca(row) if row else None

    def _encontrar_peca_por_snapshot_conn(self, conn: sqlite3.Connection, venda: VendaHistoricoItem) -> Optional[Peca]:
        """Procura uma peca ativa equivalente ao snapshot da venda."""
        row = conn.execute(
            """
            SELECT id, referencia, categoria, marca, designacao, preco, quantidade, local, extras
            FROM pecas
            WHERE referencia = ? AND categoria = ? AND marca = ? AND designacao = ?
            ORDER BY id ASC
            LIMIT 1
            """,
            (
                venda.referencia,
                venda.categoria,
                venda.marca,
                venda.designacao,
            ),
        ).fetchone()
        return self._row_to_peca(row) if row else None

    def _retirar_do_stock(self, conn: sqlite3.Connection, venda: VendaHistoricoItem, quantidade: int) -> None:
        """Retira unidades do stock atual para acomodar aumento de quantidade numa venda antiga."""
        if quantidade <= 0:
            return

        peca = None
        if venda.peca_id:
            peca = self._obter_peca_por_id_conn(conn, venda.peca_id)
        if not peca:
            peca = self._encontrar_peca_por_snapshot_conn(conn, venda)
        if not peca:
            raise ValueError("Nao existe stock disponivel para aumentar esta venda.")

        stock_atual = max(0, int(peca.quantidade))
        if quantidade > stock_atual:
            raise ValueError("A quantidade excede o stock atual disponivel para este material.")

        nova_quantidade = stock_atual - quantidade
        if nova_quantidade <= 0:
            conn.execute("DELETE FROM pecas WHERE id = ?", (peca.id,))
            return

        conn.execute("UPDATE pecas SET quantidade = ? WHERE id = ?", (nova_quantidade, peca.id))

    def _repor_ao_stock(self, conn: sqlite3.Connection, venda: VendaHistoricoItem, quantidade: int) -> Peca:
        """Repõe unidades no stock com base no snapshot guardado na venda."""
        quantidade_repor = max(1, int(quantidade))

        peca = None
        if venda.peca_id:
            peca = self._obter_peca_por_id_conn(conn, venda.peca_id)
        if not peca:
            peca = self._encontrar_peca_por_snapshot_conn(conn, venda)

        if peca:
            nova_quantidade = max(0, int(peca.quantidade)) + quantidade_repor
            conn.execute("UPDATE pecas SET quantidade = ? WHERE id = ?", (nova_quantidade, peca.id))
            return Peca(
                id=peca.id,
                referencia=peca.referencia,
                categoria=peca.categoria,
                marca=peca.marca,
                designacao=peca.designacao,
                preco=peca.preco,
                quantidade=nova_quantidade,
                local=peca.local,
                extras=peca.extras,
            )

        novo_id = str(venda.peca_id or uuid4())
        extras = venda.extras or {}
        preco_snapshot = self._safe_price(venda.preco_unitario)
        conn.execute(
            """
            INSERT INTO pecas (id, referencia, categoria, marca, designacao, preco, quantidade, local, extras)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                novo_id,
                venda.referencia,
                venda.categoria,
                venda.marca,
                venda.designacao,
                preco_snapshot,
                quantidade_repor,
                venda.local,
                json.dumps(extras, ensure_ascii=False),
            ),
        )
        return Peca(
            id=novo_id,
            referencia=venda.referencia,
            categoria=venda.categoria,
            marca=venda.marca,
            designacao=venda.designacao,
            preco=preco_snapshot,
            quantidade=quantidade_repor,
            local=venda.local,
            extras=extras,
        )

    def _criar_registo_venda(
        self,
        conn: sqlite3.Connection,
        peca: Peca,
        quantidade_vendida: int,
        preco_unitario: float,
    ) -> None:
        """Guarda um snapshot da venda para analise futura."""
        conn.execute(
            """
            INSERT INTO vendas (
                id, peca_id, referencia, categoria, marca, designacao,
                preco_unitario, quantidade_vendida, local, vendida_em, extras
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                str(uuid4()),
                peca.id,
                peca.referencia,
                peca.categoria,
                peca.marca,
                peca.designacao,
                self._safe_price(preco_unitario),
                max(1, int(quantidade_vendida)),
                peca.local,
                datetime.now().isoformat(timespec="seconds"),
                json.dumps(peca.extras or {}, ensure_ascii=False),
            ),
        )

    @staticmethod
    def _normalizar_ordem_colunas(colunas: List[ColunaSchema]) -> List[ColunaSchema]:
        """Mantem Preco imediatamente antes de Quantidade, independentemente da ordem guardada."""
        resultado = [coluna for coluna in colunas if coluna.chave not in {"preco", "quantidade"}]
        coluna_preco = next((coluna for coluna in colunas if coluna.chave == "preco"), None)
        coluna_quantidade = next((coluna for coluna in colunas if coluna.chave == "quantidade"), None)

        if coluna_preco:
            resultado.append(coluna_preco)
        if coluna_quantidade:
            resultado.append(coluna_quantidade)
        return resultado

    @staticmethod
    def _parse_extras(raw: object) -> dict[str, object]:
        """Converte JSON de extras em dicionario."""
        if raw is None:
            return {}
        if isinstance(raw, dict):
            return raw
        try:
            parsed = json.loads(str(raw))
            return parsed if isinstance(parsed, dict) else {}
        except json.JSONDecodeError:
            return {}

    def _normalizar_chave_coluna(self, value: str) -> str:
        """Transforma nome em chave segura."""
        normalized = self._normalizar_texto(value)
        return re.sub(r"[^a-z0-9]+", "_", normalized).strip("_")

    @staticmethod
    def _normalizar_texto(value: object) -> str:
        """Remove acentos e normaliza caixa."""
        raw = str(value or "").strip().lower()
        return "".join(char for char in unicodedata.normalize("NFD", raw) if unicodedata.category(char) != "Mn")

    @staticmethod
    def _safe_int(value: object) -> int:
        """Converte para inteiro de forma robusta."""
        if value is None:
            return 0
        try:
            return int(float(str(value)))
        except (TypeError, ValueError):
            return 0

    @staticmethod
    def _safe_price(value: object) -> float:
        """Converte preco para float nao negativo com duas casas decimais."""
        if value is None:
            return 0
        try:
            return round(max(0.0, float(str(value).replace(",", "."))), 2)
        except (TypeError, ValueError):
            return 0

    @staticmethod
    def _match_termo(peca: Peca, termo: str) -> bool:
        """Filtro textual para pesquisa."""
        extras_texto = " ".join(str(valor) for valor in peca.extras.values() if valor is not None)
        conjunto = " ".join(
            [
                peca.referencia,
                peca.categoria,
                peca.marca,
                peca.designacao,
                f"{peca.preco:.2f}",
                str(peca.quantidade),
                peca.local or "",
                peca.id,
                extras_texto,
            ]
        ).lower()
        return termo in conjunto

    @staticmethod
    def _formatar_mes_dashboard(chave_mes: str) -> str:
        """Converte YYYY-MM numa etiqueta curta para o dashboard."""
        try:
            data = datetime.strptime(chave_mes, "%Y-%m")
        except ValueError:
            return chave_mes

        meses = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
        return f"{meses[data.month - 1]} {data.year}"

    @staticmethod
    def _montar_etiqueta_material(referencia: object, designacao: object) -> str:
        """Gera uma etiqueta curta para graficos por material."""
        referencia_texto = str(referencia or "").strip()
        designacao_texto = str(designacao or "").strip()
        if referencia_texto and designacao_texto:
            return f"{referencia_texto} - {designacao_texto}"
        return referencia_texto or designacao_texto or "Material sem nome"
